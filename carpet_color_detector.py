import os
import numpy as np
from PIL import Image
from sklearn.cluster import KMeans
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage

# =============================================
# PATHS - CONFIGURED FOR MAC
# =============================================
IMAGE_FOLDER = "/Users/chetan.p/Desktop/G Data/AsterlaneHome/Color_Tracker/carpet_images"
OUTPUT_FILE  = "/Users/chetan.p/Desktop/G Data/AsterlaneHome/Color_Tracker/carpet_colors_result.xlsx"
THUMB_FOLDER = "/Users/chetan.p/Desktop/G Data/AsterlaneHome/Color_Tracker/thumbnails"
# =============================================

# Simple, industry-standard carpet color names
COLOR_NAMES = {
    "White":        (255, 255, 255),
    "Ivory":        (240, 234, 214),
    "Cream":        (245, 235, 210),
    "Beige":        (220, 205, 180),
    "Tan":          (195, 165, 120),
    "Light Brown":  (180, 130, 85),
    "Brown":        (130, 80, 40),
    "Dark Brown":   (80, 50, 25),
    "Light Grey":   (200, 200, 200),
    "Grey":         (150, 150, 150),
    "Dark Grey":    (90, 90, 90),
    "Black":        (20, 20, 20),
    "Red":          (190, 40, 40),
    "Dark Red":     (120, 20, 20),
    "Pink":         (220, 150, 160),
    "Orange":       (220, 130, 50),
    "Yellow":       (230, 210, 80),
    "Gold":         (200, 170, 60),
    "Light Green":  (130, 180, 120),
    "Green":        (60, 120, 60),
    "Dark Green":   (30, 70, 40),
    "Light Blue":   (140, 180, 210),
    "Blue":         (60, 100, 160),
    "Dark Blue":    (25, 40, 90),
    "Navy":         (20, 30, 70),
    "Purple":       (110, 60, 130),
    "Light Purple": (170, 140, 190),
}

def get_color_name(rgb):
    min_dist = float('inf')
    closest = "Unknown"
    for name, ref in COLOR_NAMES.items():
        dist = sum((int(a) - int(b)) ** 2 for a, b in zip(rgb, ref)) ** 0.5
        if dist < min_dist:
            min_dist = dist
            closest = name
    return closest

def get_hex(rgb):
    return "#{:02x}{:02x}{:02x}".format(int(rgb[0]), int(rgb[1]), int(rgb[2]))

def extract_colors(image_path, n_colors=3):
    try:
        img = Image.open(image_path).convert("RGB").resize((500, 500))
        pixels = np.array(img).reshape(-1, 3)

        kmeans = KMeans(n_clusters=n_colors, random_state=42, n_init=10)
        kmeans.fit(pixels)

        unique_labels, counts = np.unique(kmeans.labels_, return_counts=True)
        sorted_indices = np.argsort(-counts)
        total = len(kmeans.labels_)

        colors = []
        for idx in sorted_indices:
            rgb = kmeans.cluster_centers_[idx].astype(int)
            colors.append({
                "name":       get_color_name(rgb),
                "hex":        get_hex(rgb),
                "rgb":        (int(rgb[0]), int(rgb[1]), int(rgb[2])),
                "confidence": round((counts[idx] / total) * 100, 1)
            })
        return colors
    except Exception as e:
        print(f"    ⚠️  Error: {e}")
        return [{"name": "Error", "hex": "#000000", "rgb": (0,0,0), "confidence": 0.0}] * n_colors

def create_thumbnail(image_path, thumb_path, size=(120, 150)):
    """Create a small thumbnail and save it."""
    try:
        img = Image.open(image_path).convert("RGB")
        img.thumbnail(size, Image.LANCZOS)
        img.save(thumb_path)
        return True
    except Exception as e:
        print(f"    ⚠️  Thumbnail error: {e}")
        return False

def style_excel(ws, row_count):
    header_fill  = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font    = Font(name="Calibri", size=10, color="2C2C2C")
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    border       = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )
    alt_fill = PatternFill(start_color="F5F5F8", end_color="F5F5F8", fill_type="solid")

    widths = {"A":6, "B":18, "C":40, "D":7, "E":16, "F":10, "G":11,
              "H":7, "I":16, "J":10, "K":11,
              "L":7, "M":16, "N":10, "O":11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border

    swatch_cols  = {4, 8, 12}
    thumb_col    = 2

    for row_idx in range(2, row_count + 2):
        ws.row_dimensions[row_idx].height = 100
        for col_idx in range(1, 16):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font      = data_font
            cell.border    = border
            cell.alignment = left if col_idx == 3 else center
            if row_idx % 2 == 0 and col_idx not in swatch_cols and col_idx != thumb_col:
                cell.fill = alt_fill

    ws.freeze_panes = "A2"

def main():
    print("=" * 55)
    print("   CARPET COLOR DETECTOR v4 — Mac Edition")
    print("=" * 55)
    print(f"\n📂 Folder : {IMAGE_FOLDER}\n")

    os.makedirs(THUMB_FOLDER, exist_ok=True)

    valid = ('.jpg','.jpeg','.png','.bmp','.tiff','.webp')
    files = sorted([f for f in os.listdir(IMAGE_FOLDER) if f.lower().endswith(valid)])

    if not files:
        print("❌ No images found! Check folder path.")
        return

    print(f"📷 Found {len(files)} images. Processing...\n")

    rows     = []
    swatches = []
    thumbs   = []

    for i, fname in enumerate(files, 1):
        img_path   = os.path.join(IMAGE_FOLDER, fname)
        thumb_path = os.path.join(THUMB_FOLDER, f"thumb_{i}_{fname.split('.')[0]}.jpg")

        create_thumbnail(img_path, thumb_path)
        thumbs.append(thumb_path)

        colors = extract_colors(img_path)

        rows.append([
            i, "",  fname,
            "", colors[0]["name"], f"{colors[0]['confidence']}%", colors[0]["hex"],
            "", colors[1]["name"], f"{colors[1]['confidence']}%", colors[1]["hex"],
            "", colors[2]["name"], f"{colors[2]['confidence']}%", colors[2]["hex"],
        ])
        swatches.append((colors[0]["rgb"], colors[1]["rgb"], colors[2]["rgb"]))

        print(f"  ✓ [{i}/{len(files)}] {fname}")
        print(f"      Primary   → {colors[0]['name']:14s} | {colors[0]['confidence']:>5}% | {colors[0]['hex']}")
        print(f"      Secondary → {colors[1]['name']:14s} | {colors[1]['confidence']:>5}% | {colors[1]['hex']}")
        print(f"      Tertiary  → {colors[2]['name']:14s} | {colors[2]['confidence']:>5}% | {colors[2]['hex']}\n")

    wb = Workbook()
    ws = wb.active
    ws.title = "Carpet Colors"

    ws.append(["S.No", "Thumbnail", "Image Name",
               "Swatch", "Primary Color",   "Primary %",   "Primary Hex",
               "Swatch", "Secondary Color", "Secondary %", "Secondary Hex",
               "Swatch", "Tertiary Color",  "Tertiary %",  "Tertiary Hex"])

    for row in rows:
        ws.append(row)

    for idx, thumb_path in enumerate(thumbs):
        excel_row = idx + 2
        if os.path.exists(thumb_path):
            img = XLImage(thumb_path)
            img.width  = 100
            img.height = 90
            cell_ref = f"B{excel_row}"
            img.anchor = cell_ref
            ws.add_image(img)

    for idx, (pri, sec, ter) in enumerate(swatches):
        r_idx = idx + 2
        for col, rgb in [(4, pri), (8, sec), (12, ter)]:
            r, g, b = rgb
            ws.cell(row=r_idx, column=col).fill = PatternFill(
                start_color=f"{r:02x}{g:02x}{b:02x}",
                end_color  =f"{r:02x}{g:02x}{b:02x}",
                fill_type  ="solid"
            )

    style_excel(ws, len(rows))
    wb.save(OUTPUT_FILE)

    print("=" * 55)
    print(f"  ✅ DONE!")
    print(f"  📄 {OUTPUT_FILE}")
    print("=" * 55)

if __name__ == "__main__":
    main()
